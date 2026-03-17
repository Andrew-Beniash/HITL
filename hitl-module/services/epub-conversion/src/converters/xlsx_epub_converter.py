"""XlsxEpubConverter — XLSX → EPUB3 using openpyxl + ebooklib.

Each visible worksheet becomes one or more EpubHtml chapters (paginated when
the sheet exceeds PAGINATION_THRESHOLD rows).  Merged cells produce correct
colspan/rowspan; frozen rows become <thead>; formula cells are evaluated via a
two-stage cache (openpyxl data_only → formulas library).  Charts are embedded
as 400×200 px Pillow placeholder PNGs.
"""

from __future__ import annotations

import hashlib
import html as html_mod
import io
import pathlib
import re
import tempfile
from datetime import datetime, timedelta, timezone
from typing import Any

import openpyxl
from ebooklib import epub
from openpyxl.utils.cell import coordinate_to_tuple
from PIL import Image, ImageDraw

from src.converters.base import BaseConverter, ConversionError


class XlsxEpubConverter(BaseConverter):
    MAX_ROWS_PER_PAGE = 500
    PAGINATION_THRESHOLD = 5000

    def __init__(self) -> None:
        self._source_path: str = ""
        self._formula_cache: dict[str, Any] = {}

    # ── Public entry point ────────────────────────────────────────────────────

    def convert(self, source_path: str) -> tuple[bytes, dict]:
        self._source_path = source_path
        source_bytes = pathlib.Path(source_path).read_bytes()
        source_hash = hashlib.sha256(source_bytes).hexdigest()

        wb = openpyxl.load_workbook(source_path, data_only=False)
        book = epub.EpubBook()
        book.set_title(pathlib.Path(source_path).stem)
        book.set_language("en")

        chapters: list[epub.EpubHtml] = []
        sheets_manifest: list[dict] = []
        degradation_notices: list[dict] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.sheet_state == "hidden":
                continue

            sheet_manifest, notices, chapter_items = self._convert_sheet(
                ws, sheet_name, book
            )
            sheets_manifest.append(sheet_manifest)
            degradation_notices.extend(notices)
            chapters.extend(chapter_items)

        # Build NCX/Nav TOC
        book.toc = [  # type: ignore[assignment]
            (
                epub.Section(s["name"]),
                [c for c in chapters if c.sheet == s["name"]],  # type: ignore[attr-defined]
            )
            for s in sheets_manifest
        ]
        book.add_item(epub.EpubNcx())
        book.add_item(epub.EpubNav())

        epub_bytes = self._package(book, chapters)
        manifest = {
            "sourceFormat": "xlsx",
            "sourceFileHash": source_hash,
            "convertedAt": datetime.now(timezone.utc).isoformat(),
            "sheets": sheets_manifest,
            "degradationNotices": degradation_notices,
        }
        return epub_bytes, manifest

    # ── Sheet converter ───────────────────────────────────────────────────────

    def _convert_sheet(
        self,
        ws: Any,
        sheet_name: str,
        book: epub.EpubBook,
    ) -> tuple[dict, list[dict], list[epub.EpubHtml]]:
        notices: list[dict] = []
        chapter_items: list[epub.EpubHtml] = []

        # ── Merge map ────────────────────────────────────────────────────────
        # { (row, col): (rowspan, colspan) } for master cells
        # { (row, col): None }               for covered cells
        merge_map: dict[tuple[int, int], tuple[int, int] | None] = {}
        for rng in ws.merged_cells.ranges:
            min_r, min_c = rng.min_row, rng.min_col
            merge_map[(min_r, min_c)] = (
                rng.max_row - min_r + 1,
                rng.max_col - min_c + 1,
            )
            for r, c in rng.cells:
                if (r, c) != (min_r, min_c):
                    merge_map[(r, c)] = None

        # ── Frozen-row detection ─────────────────────────────────────────────
        header_count = 0
        if ws.freeze_panes:
            freeze_row, _ = coordinate_to_tuple(str(ws.freeze_panes))
            header_count = max(0, freeze_row - 1)

        # ── Degradation notices ───────────────────────────────────────────────
        if ws.conditional_formatting:
            notices.append(
                {
                    "sheet": sheet_name,
                    "type": "conditional_formatting_omitted",
                    "message": (
                        f"Sheet '{sheet_name}' uses conditional formatting "
                        "which is not rendered in EPUB output."
                    ),
                }
            )

        # ── Formula cache (built lazily — only when formula cells exist) ─────
        self._formula_cache = self._build_formula_cache(
            self._source_path, sheet_name, ws
        )

        # ── Chart placeholders ────────────────────────────────────────────────
        chart_items: list[dict] = []
        for chart_id, chart in enumerate(getattr(ws, "_charts", [])):
            chart_items.append(
                self._extract_chart(book, chart, chart_id, sheet_name)
            )

        # ── Rows ──────────────────────────────────────────────────────────────
        rows = list(ws.iter_rows())
        header_rows = rows[:header_count]
        data_rows = rows[header_count:]

        # ── Pagination ────────────────────────────────────────────────────────
        if len(data_rows) > self.PAGINATION_THRESHOLD:
            pages = [
                data_rows[i : i + self.MAX_ROWS_PER_PAGE]
                for i in range(0, len(data_rows), self.MAX_ROWS_PER_PAGE)
            ]
        else:
            pages = [data_rows]

        safe_name = re.sub(r"[^a-zA-Z0-9_-]", "_", sheet_name)

        for page_idx, page_rows in enumerate(pages):
            page_title = (
                sheet_name
                if len(pages) == 1
                else f"{sheet_name} ({page_idx + 1}/{len(pages)})"
            )
            html_content = self._rows_to_html(
                merge_map,
                header_rows,
                page_rows,
                page_idx,
                len(pages),
                chart_items if page_idx == 0 else [],
                page_title,
            )
            chapter = epub.EpubHtml(
                title=page_title,
                file_name=f"sheet_{safe_name}_{page_idx}.xhtml",
                lang="en",
            )
            chapter.content = html_content.encode("utf-8")
            chapter.sheet = sheet_name  # type: ignore[attr-defined]
            book.add_item(chapter)
            chapter_items.append(chapter)

        sheet_manifest = {
            "name": sheet_name,
            "rows": len(rows),
            "pages": len(pages),
        }
        return sheet_manifest, notices, chapter_items

    # ── HTML table builder ────────────────────────────────────────────────────

    def _rows_to_html(
        self,
        merge_map: dict,
        header_rows: list,
        data_rows: list,
        page_idx: int,
        total_pages: int,
        chart_items: list[dict],
        page_title: str,
    ) -> str:
        parts: list[str] = [
            "<?xml version='1.0' encoding='utf-8'?>",
            "<!DOCTYPE html>",
            '<html xmlns="http://www.w3.org/1999/xhtml">',
            "<head>",
            f"  <title>{html_mod.escape(page_title)}</title>",
            '  <link rel="stylesheet" href="../styles/platform.css" type="text/css"/>',
            "</head>",
            "<body>",
        ]

        pagination_html = ""
        if total_pages > 1:
            pagination_html = (
                f'<div class="sheet-pagination">'
                f"Page {page_idx + 1} of {total_pages}"
                f"</div>"
            )
            parts.append(pagination_html)

        for ci in chart_items:
            parts.append(
                f'<figure>'
                f'<img src="../images/{ci["filename"]}" '
                f'alt="{html_mod.escape(ci["title"])}"/>'
                f'<figcaption>{html_mod.escape(ci["title"])}</figcaption>'
                f'</figure>'
            )

        parts.append('<table class="sheet-table">')

        if header_rows:
            parts.append("<thead>")
            for row in header_rows:
                parts.append("<tr>")
                for cell in row:
                    pos = (cell.row, cell.column)
                    if pos in merge_map and merge_map[pos] is None:
                        continue
                    merge_entry = merge_map.get(pos)
                    attrs = self._cell_attrs(cell, merge_entry)
                    content = self._apply_inline_formatting(
                        cell, self._format_cell(cell)
                    )
                    parts.append(f"  <th{attrs}>{content}</th>")
                parts.append("</tr>")
            parts.append("</thead>")

        parts.append("<tbody>")
        for row in data_rows:
            parts.append("<tr>")
            for cell in row:
                pos = (cell.row, cell.column)
                if pos in merge_map and merge_map[pos] is None:
                    continue
                merge_entry = merge_map.get(pos)
                attrs = self._cell_attrs(cell, merge_entry)
                content = self._apply_inline_formatting(
                    cell, self._format_cell(cell)
                )
                parts.append(f"  <td{attrs}>{content}</td>")
            parts.append("</tr>")
        parts.append("</tbody>")
        parts.append("</table>")

        if total_pages > 1:
            parts.append(pagination_html)

        parts.append("</body>")
        parts.append("</html>")
        return "\n".join(parts)

    # ── Cell helpers ──────────────────────────────────────────────────────────

    def _cell_attrs(
        self,
        cell: Any,
        merge_entry: tuple[int, int] | None,
    ) -> str:
        """Return attribute string (leading space) for a <td>/<th> element."""
        attrs: list[str] = [
            f'data-row="{cell.row}"',
            f'data-col="{cell.column}"',
        ]

        if isinstance(merge_entry, tuple):
            rowspan, colspan = merge_entry
            if rowspan > 1:
                attrs.append(f'rowspan="{rowspan}"')
            if colspan > 1:
                attrs.append(f'colspan="{colspan}"')

        css: list[str] = []
        try:
            align = cell.alignment.horizontal
            if align in ("left", "center", "right", "justify"):
                css.append(f"text-align:{align}")
        except Exception:
            pass

        try:
            rgb = cell.fill.fgColor.rgb
            if rgb and rgb not in ("00000000", "FF000000"):
                css.append(f"background-color:#{rgb[2:]}")
        except Exception:
            pass

        if css:
            attrs.append(f'style="{"; ".join(css)}"')

        return " " + " ".join(attrs)

    def _apply_inline_formatting(self, cell: Any, text: str) -> str:
        """Wrap *text* in <strong>/<em> per cell font."""
        try:
            bold = cell.font.bold
            italic = cell.font.italic
        except Exception:
            bold = italic = False

        escaped = html_mod.escape(text)
        if italic:
            escaped = f"<em>{escaped}</em>"
        if bold:
            escaped = f"<strong>{escaped}</strong>"
        return escaped

    def _format_cell(self, cell: Any) -> str:
        """Return the display-formatted string value for *cell*."""
        if cell.value is None:
            return ""

        # Formula cell: look up evaluated value from cache
        if cell.data_type == "f":
            cached = self._formula_cache.get(cell.coordinate)
            if cached is None:
                return "#EVAL_ERROR"
            return self._apply_number_format(cached, cell.number_format)

        return self._apply_number_format(cell.value, cell.number_format)

    def _apply_number_format(self, value: Any, number_format: str | None) -> str:
        """Apply an Excel number-format string to a Python value."""
        if value is None:
            return ""

        # Unwrap numpy scalars / 0-d arrays returned by formulas library
        if hasattr(value, "flat"):
            value = value.flat[0]
        if hasattr(value, "item"):
            value = value.item()

        fmt = number_format or "General"
        fmt_upper = fmt.upper()

        # Date / time (contains date pattern tokens)
        if any(tok in fmt_upper for tok in ("YY", "DD-MMM", "MM/DD", "D/M")):
            try:
                if isinstance(value, (int, float)):
                    dt = datetime(1899, 12, 30) + timedelta(days=float(value))
                    return dt.strftime("%Y-%m-%d")
                import datetime as _dt
                if isinstance(value, (_dt.date, _dt.datetime)):
                    return value.strftime("%Y-%m-%d")
            except Exception:
                pass
            return str(value)

        # Percentage
        if "%" in fmt:
            try:
                return f"{float(value) * 100:.1f}%"
            except (TypeError, ValueError):
                return str(value)

        # Currency
        if "$" in fmt:
            try:
                return f"${float(value):,.2f}"
            except (TypeError, ValueError):
                return str(value)
        if "€" in fmt:
            try:
                return f"€{float(value):,.2f}"
            except (TypeError, ValueError):
                return str(value)

        # Decimal precision (count zeros after the decimal point)
        if "." in fmt:
            after_dot = fmt.split(".")[1]
            precision = sum(1 for c in after_dot if c == "0")
            try:
                return f"{float(value):.{precision}f}"
            except (TypeError, ValueError):
                return str(value)

        return str(value)

    # ── Formula cache ─────────────────────────────────────────────────────────

    def _build_formula_cache(
        self, source_path: str, ws_name: str, ws: Any
    ) -> dict[str, Any]:
        """Build {coordinate: evaluated_value} for all formula cells in *ws*.

        Stage 1: openpyxl data_only reads Excel-cached values from the XML
                 (works for workbooks saved by Excel or LibreOffice).
        Stage 2: formulas.Parser evaluates uncached formula cells by resolving
                 their cell-reference inputs against the known scalar values in
                 the same sheet (works for openpyxl-created test fixtures).
        Returns an empty dict immediately when the sheet has no formula cells.
        """
        # Fast path: skip entirely if no formula cells exist
        has_formula = any(
            cell.data_type == "f"
            for row in ws.iter_rows()
            for cell in row
        )
        if not has_formula:
            return {}

        cache: dict[str, Any] = {}

        # Stage 1 — openpyxl data_only
        try:
            wb_data = openpyxl.load_workbook(source_path, data_only=True)
            ws_data = wb_data[ws_name]
            for row in ws_data.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cache[cell.coordinate] = cell.value
            wb_data.close()
        except Exception:
            pass

        # Stage 2 — formulas.Parser with direct cell-value substitution
        # Build a flat {coordinate: scalar} map of all non-formula cells in the
        # sheet so the Parser can resolve references like A1, B3, etc.
        scalar_values: dict[str, Any] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type != "f" and cell.value is not None:
                    scalar_values[cell.coordinate] = cell.value

        # Include already-cached formula values (dependency chain resolution)
        scalar_values.update(cache)

        try:
            import formulas as fm  # noqa: PLC0415
            import numpy as np  # noqa: PLC0415

            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type != "f" or cell.coordinate in cache:
                        continue
                    formula_str = str(cell.value)  # e.g. "=A1+A2"
                    try:
                        func = fm.Parser().ast(formula_str)[1].compile()
                        inputs: dict[str, Any] = {}
                        all_resolved = True
                        for inp_name in func.inputs:
                            ref_val = scalar_values.get(inp_name.upper())
                            if ref_val is None:
                                all_resolved = False
                                break
                            try:
                                inputs[inp_name] = np.array([[float(ref_val)]])
                            except (TypeError, ValueError):
                                inputs[inp_name] = np.array([[ref_val]])

                        if not all_resolved or not inputs:
                            continue

                        result = func(**inputs)
                        if hasattr(result, "flat"):
                            val = result.flat[0]
                        else:
                            val = result
                        if hasattr(val, "item"):
                            val = val.item()
                        if val is not None:
                            cache[cell.coordinate] = val
                            scalar_values[cell.coordinate] = val
                    except Exception:
                        pass
        except ImportError:
            pass

        return cache

    # ── Chart extraction ──────────────────────────────────────────────────────

    def _extract_chart(
        self,
        book: epub.EpubBook,
        chart: Any,
        chart_id: int,
        sheet_name: str,
    ) -> dict:
        """Rasterise a chart to a 400×200 px PNG placeholder via Pillow."""
        title = f"Chart {chart_id + 1}"
        try:
            t = chart.title
            if t is not None:
                if hasattr(t, "text") and t.text:
                    if hasattr(t.text, "rich"):
                        title = "".join(
                            getattr(r, "text", "") for r in t.text.rich
                        )
                    elif hasattr(t.text, "value"):
                        title = str(t.text.value)
                    else:
                        title = str(t.text)
                elif hasattr(t, "value"):
                    title = str(t.value)
                else:
                    title = str(t) if t else title
        except Exception:
            pass

        img = Image.new("RGB", (400, 200), color="white")
        draw = ImageDraw.Draw(img)
        draw.rectangle([2, 2, 397, 197], outline="#cccccc", width=2)
        draw.text((10, 10), title, fill="#333333")
        draw.text((10, 80), "[Chart — static placeholder]", fill="#999999")
        buf = io.BytesIO()
        img.save(buf, format="PNG")

        safe_sheet = re.sub(r"[^a-zA-Z0-9_-]", "_", sheet_name)
        filename = f"chart_{safe_sheet}_{chart_id}.png"
        epub_img = epub.EpubImage()
        epub_img.file_name = f"images/{filename}"
        epub_img.media_type = "image/png"
        epub_img.content = buf.getvalue()
        book.add_item(epub_img)

        return {"filename": filename, "title": title}

    # ── EPUB packaging ────────────────────────────────────────────────────────

    def _package(
        self, book: epub.EpubBook, chapters: list[epub.EpubHtml]
    ) -> bytes:
        book.spine = ["nav"] + chapters
        with tempfile.NamedTemporaryFile(suffix=".epub", delete=False) as f:
            tmp_path = f.name
        try:
            epub.write_epub(tmp_path, book)
            return pathlib.Path(tmp_path).read_bytes()
        finally:
            pathlib.Path(tmp_path).unlink(missing_ok=True)
