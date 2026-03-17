"""Tests for XlsxEpubConverter — all fixtures created programmatically."""

from __future__ import annotations

import io
import time
import zipfile

import openpyxl
import pytest

from src.converters.xlsx_epub_converter import XlsxEpubConverter


def _chapter_files(epub_bytes: bytes) -> list[str]:
    """Return names of sheet chapter xhtml files inside the EPUB ZIP."""
    with zipfile.ZipFile(io.BytesIO(epub_bytes)) as zf:
        return [n for n in zf.namelist() if "sheet_" in n and n.endswith(".xhtml")]


def _chapter_content(epub_bytes: bytes, index: int = 0) -> str:
    """Read the text of the Nth chapter xhtml."""
    with zipfile.ZipFile(io.BytesIO(epub_bytes)) as zf:
        files = sorted(
            n for n in zf.namelist() if "sheet_" in n and n.endswith(".xhtml")
        )
        return zf.read(files[index]).decode()


# ── test_three_sheets ─────────────────────────────────────────────────────────

def test_three_sheets(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Alpha"
    wb.create_sheet("Beta")
    wb.create_sheet("Gamma")
    path = tmp_path / "three.xlsx"
    wb.save(str(path))

    epub_bytes, manifest = XlsxEpubConverter().convert(str(path))

    # Manifest lists three sheets
    assert len(manifest["sheets"]) == 3
    names = {s["name"] for s in manifest["sheets"]}
    assert names == {"Alpha", "Beta", "Gamma"}

    # Three XHTML chapter files in the EPUB ZIP
    chapters = _chapter_files(epub_bytes)
    assert len(chapters) == 3

    # NCX/nav is present
    with zipfile.ZipFile(io.BytesIO(epub_bytes)) as zf:
        all_files = zf.namelist()
    assert any("toc.ncx" in f for f in all_files)


# ── test_frozen_header ────────────────────────────────────────────────────────

def test_frozen_header(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Score"
    ws["A2"] = "Alice"
    ws["B2"] = 95
    ws["A3"] = "Bob"
    ws["B3"] = 87
    ws.freeze_panes = "A2"  # row 1 is the frozen header
    path = tmp_path / "frozen.xlsx"
    wb.save(str(path))

    epub_bytes, _ = XlsxEpubConverter().convert(str(path))
    content = _chapter_content(epub_bytes)

    assert "<thead>" in content, "Frozen row must produce a <thead> element"
    assert "<tbody>" in content, "Data rows must produce a <tbody> element"


# ── test_merged_cells ─────────────────────────────────────────────────────────

def test_merged_cells(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Merged"
    ws["A1"] = "Merged Header"
    ws.merge_cells("A1:B1")  # colspan=2
    ws["A2"] = "Left"
    ws["B2"] = "Right"
    path = tmp_path / "merged.xlsx"
    wb.save(str(path))

    epub_bytes, _ = XlsxEpubConverter().convert(str(path))
    content = _chapter_content(epub_bytes)

    assert 'colspan="2"' in content, "Merged A1:B1 must produce colspan=2"
    # B1 (covered cell) must NOT appear as a separate td/th
    b1_occurrences = content.count('data-col="2"')
    # Only A2's col=2 cell (B2) should appear, not the covered B1
    assert b1_occurrences <= 1, "Covered cell B1 must be skipped"


# ── test_pagination ───────────────────────────────────────────────────────────

def test_pagination(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BigSheet"
    for row_idx in range(1, 6001):  # exactly 6 000 rows → 12 pages of 500
        ws.cell(row=row_idx, column=1, value=row_idx)
    path = tmp_path / "large.xlsx"
    wb.save(str(path))

    epub_bytes, manifest = XlsxEpubConverter().convert(str(path))
    chapters = _chapter_files(epub_bytes)

    assert len(chapters) == 12, (
        f"6 000 rows / 500 per page = 12 chapters; got {len(chapters)}"
    )
    assert manifest["sheets"][0]["pages"] == 12

    # Pagination labels present in the content
    content = _chapter_content(epub_bytes, index=0)
    assert "Page 1 of 12" in content


# ── test_formula_cell ─────────────────────────────────────────────────────────

def test_formula_cell(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulas"
    ws["A1"] = 5
    ws["A2"] = 10
    ws["A3"] = "=A1+A2"  # should evaluate to 15
    path = tmp_path / "formula.xlsx"
    wb.save(str(path))

    epub_bytes, _ = XlsxEpubConverter().convert(str(path))
    content = _chapter_content(epub_bytes)

    assert "=A1+A2" not in content, "Formula expression must not appear in EPUB"
    assert "15" in content, "Evaluated result (15) must appear in EPUB"


# ── test_conditional_formatting ───────────────────────────────────────────────

def test_conditional_formatting(tmp_path):
    from openpyxl.formatting.rule import ColorScaleRule

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CF"
    for i, val in enumerate([1, 50, 100], start=1):
        ws.cell(row=i, column=1, value=val)
    rule = ColorScaleRule(
        start_type="min",
        start_color="FF0000",
        end_type="max",
        end_color="00FF00",
    )
    ws.conditional_formatting.add("A1:A3", rule)
    path = tmp_path / "cf.xlsx"
    wb.save(str(path))

    _, manifest = XlsxEpubConverter().convert(str(path))

    notice_types = [n["type"] for n in manifest["degradationNotices"]]
    assert "conditional_formatting_omitted" in notice_types, (
        "Conditional formatting must produce a degradation notice"
    )


# ── test_performance ──────────────────────────────────────────────────────────

def test_performance(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PerfSheet"
    for r in range(1, 5001):
        for c in range(1, 6):
            ws.cell(row=r, column=c, value=r * c)
    path = tmp_path / "perf.xlsx"
    wb.save(str(path))

    start = time.time()
    XlsxEpubConverter().convert(str(path))
    elapsed = time.time() - start

    assert elapsed < 10.0, (
        f"5 000-row workbook conversion took {elapsed:.1f}s — must be under 10s"
    )
